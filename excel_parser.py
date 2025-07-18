import pandas as pd
import streamlit as st
from typing import Dict, List, Any, Optional, Tuple
import io
import re
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import gc  # 垃圾回收

class ExcelParser:
    def __init__(self):
        # 缓存设置 - 内存优化
        self.sheet_cache = {}
        self.max_cache_size = 20  # 减少缓存大小
        self.cache_timeout = 180  # 3分钟缓存
        
        # 性能设置 - 快速解析优化
        self.max_rows_to_scan = 1000   # 大幅减少扫描行数
        self.preview_rows = 5          # 减少预览行数
        self.max_cols_to_check = 20    # 最大检查列数
        self.max_data_check_rows = 5   # 只检查前5行判断是否有数据
        
        # 文件大小限制
        self.max_file_size = 10 * 1024 * 1024  # 10MB
    
    def validate_excel_file(self, file_content: bytes) -> Tuple[bool, str]:
        """验证Excel文件格式 - 内存优化版本"""
        try:
            # 文件大小检查
            if len(file_content) > self.max_file_size:
                return False, f"文件大小超过限制 ({self.max_file_size // (1024*1024)}MB)"
            
            # 文件格式快速验证
            if len(file_content) < 1024:  # 太小的文件
                return False, "文件太小，可能不是有效的Excel文件"
            
            # 尝试快速加载文件头
            try:
                # 使用内存优化的只读模式
                workbook = load_workbook(
                    io.BytesIO(file_content), 
                    read_only=True,      # 只读模式
                    data_only=True,      # 只读数据值
                    keep_links=False     # 不保留外部链接
                )
                
                # 基本检查
                if len(workbook.sheetnames) == 0:
                    workbook.close()
                    return False, "Excel文件中没有工作表"
                
                workbook.close()
                return True, "文件格式验证通过"
                
            except openpyxl.utils.exceptions.InvalidFileException:
                return False, "不是有效的Excel文件格式"
            except Exception as e:
                return False, f"文件格式验证失败: {str(e)}"
                
        except Exception as e:
            return False, f"文件验证过程中出错: {str(e)}"
    
    def get_file_statistics(self, file_content: bytes, progress_callback=None) -> Dict[str, Any]:
        """获取Excel文件统计信息 - 快速解析优化版本"""
        try:
            if progress_callback:
                progress_callback(10, "正在打开Excel文件...")
            
            # 内存优化的文件加载
            workbook = load_workbook(
                io.BytesIO(file_content), 
                read_only=True,
                data_only=True,
                keep_links=False,
                keep_vba=False  # 不保留VBA宏
            )
            
            sheet_names = workbook.sheetnames
            total_sheets = len(sheet_names)
            file_size = len(file_content)
            
            if progress_callback:
                progress_callback(30, f"发现 {total_sheets} 个工作表，开始快速分析...")
            
            sheets_info = []
            valid_sheet_names = []
            processed_count = 0
            
            for sheet_name in sheet_names:
                try:
                    if progress_callback:
                        progress = 30 + int((processed_count / total_sheets) * 60)
                        progress_callback(progress, f"正在分析: {sheet_name}")
                    
                    # 快速获取工作表基本信息
                    worksheet = workbook[sheet_name]
                    
                    # 快速获取维度（不扫描全表）
                    max_row = worksheet.max_row or 0
                    max_col = worksheet.max_column or 0
                    
                    # 快速数据检查 - 只检查前几行
                    has_data = self._quick_data_check(worksheet, max_row, max_col)
                    
                    sheet_info = {
                        'name': sheet_name,
                        'rows': max_row,
                        'columns': max_col,
                        'has_data': has_data
                    }
                    
                    sheets_info.append(sheet_info)
                    
                    # 只有有数据的工作表才算作有效门店
                    if has_data:
                        valid_sheet_names.append(sheet_name)
                    
                    processed_count += 1
                        
                except Exception as e:
                    st.warning(f"分析工作表 '{sheet_name}' 时出错: {str(e)}")
                    sheets_info.append({
                        'name': sheet_name,
                        'rows': 0,
                        'columns': 0,
                        'has_data': False
                    })
                    processed_count += 1
            
            workbook.close()
            
            # 强制垃圾回收
            gc.collect()
            
            if progress_callback:
                progress_callback(100, f"分析完成！发现 {len(valid_sheet_names)} 个有效门店")
            
            return {
                'total_sheets': total_sheets,
                'file_size': file_size,
                'sheets_info': sheets_info,
                'sheet_names': valid_sheet_names,
                'valid_sheets': len(valid_sheet_names)
            }
            
        except Exception as e:
            if progress_callback:
                progress_callback(0, f"分析失败: {str(e)}")
            st.error(f"分析Excel文件失败: {str(e)}")
            return {
                'total_sheets': 0,
                'file_size': len(file_content),
                'sheets_info': [],
                'sheet_names': [],
                'valid_sheets': 0
            }
    
    def _quick_data_check(self, worksheet, max_row: int, max_col: int) -> bool:
        """快速检查工作表是否有数据 - 性能优化版本"""
        try:
            # 基本检查
            if max_row <= 1 or max_col <= 0:
                return False
            
            # 限制检查范围以提高性能
            check_rows = min(max_row, self.max_data_check_rows + 1)  # +1 因为要跳过标题行
            check_cols = min(max_col, self.max_cols_to_check)
            
            # 检查前几行是否有实际数据（跳过第一行标题）
            data_found = False
            for row_num in range(2, check_rows + 1):  # 从第2行开始
                for col_num in range(1, check_cols + 1):
                    try:
                        cell_value = worksheet.cell(row_num, col_num).value
                        if cell_value is not None:
                            # 转换为字符串并检查是否有实际内容
                            str_value = str(cell_value).strip()
                            if str_value and str_value.lower() not in ['none', 'null', '']:
                                data_found = True
                                break
                    except Exception:
                        # 如果读取单元格出错，继续检查下一个
                        continue
                
                if data_found:
                    break
            
            return data_found
            
        except Exception:
            # 如果快速检查失败，保守地返回 True
            return max_row > 1 and max_col > 0
    
    def get_sheet_names_fast(self, file_content: bytes) -> List[str]:
        """超快速获取有效工作表名称"""
        try:
            # 先检查缓存
            file_hash = str(hash(file_content))
            cache_key = f"sheet_names_{file_hash}"
            
            cached_names = self._get_from_cache(cache_key)
            if cached_names:
                return cached_names
            
            # 快速获取统计信息
            stats = self.get_file_statistics(file_content)
            sheet_names = stats.get('sheet_names', [])
            
            # 缓存结果
            self._add_to_cache(cache_key, sheet_names)
            
            return sheet_names
        except Exception as e:
            st.error(f"获取工作表名称失败: {str(e)}")
            return []
    
    def get_sheet_preview(self, file_content: bytes, sheet_name: str, limit: int = 5) -> Optional[Dict[str, Any]]:
        """获取工作表预览数据 - 内存优化版本"""
        try:
            # 检查缓存
            cache_key = f"{sheet_name}_preview_{limit}"
            cached_data = self._get_from_cache(cache_key)
            if cached_data:
                return cached_data
            
            # 内存优化的文件读取
            workbook = load_workbook(
                io.BytesIO(file_content), 
                read_only=True,
                data_only=True,
                keep_links=False
            )
            
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return None
            
            worksheet = workbook[sheet_name]
            
            # 获取基本信息
            total_rows = worksheet.max_row or 0
            total_columns = worksheet.max_column or 0
            
            # 读取预览数据 - 限制范围
            preview_data = []
            headers = []
            
            # 读取表头（第一行）- 限制列数
            if total_rows > 0:
                max_preview_cols = min(total_columns, self.max_cols_to_check)
                for col_num in range(1, max_preview_cols + 1):
                    try:
                        cell_value = worksheet.cell(1, col_num).value
                        header = str(cell_value).strip() if cell_value is not None else f"列{col_num}"
                        headers.append(header)
                    except Exception:
                        headers.append(f"列{col_num}")
            
            # 读取数据行 - 限制行数
            max_preview_rows = min(total_rows, limit + 1)
            for row_num in range(2, max_preview_rows + 1):  # 跳过表头
                row_data = {}
                has_data = False
                
                for col_idx, header in enumerate(headers):
                    col_num = col_idx + 1
                    try:
                        cell_value = worksheet.cell(row_num, col_num).value
                        value = str(cell_value).strip() if cell_value is not None else ""
                        row_data[header] = value
                        if value:
                            has_data = True
                    except Exception:
                        row_data[header] = ""
                
                if has_data:
                    preview_data.append(row_data)
                
                # 如果已经获得足够的预览数据，提前退出
                if len(preview_data) >= limit:
                    break
            
            workbook.close()
            gc.collect()  # 强制垃圾回收
            
            result = {
                'total_rows': total_rows,
                'total_columns': total_columns,
                'preview_data': preview_data,
                'headers': headers
            }
            
            # 缓存结果
            self._add_to_cache(cache_key, result)
            
            return result
            
        except Exception as e:
            st.error(f"获取工作表预览失败: {str(e)}")
            return None
    
    def search_in_sheet(self, file_content: bytes, sheet_name: str, search_code: str, fuzzy_match: bool = True) -> Optional[Dict[str, Any]]:
        """在指定工作表中搜索编码 - 内存优化版本"""
        try:
            # 检查缓存
            cache_key = f"{sheet_name}_data"
            cached_data = self._get_from_cache(cache_key)
            
            if not cached_data:
                # 内存优化的文件读取
                workbook = load_workbook(
                    io.BytesIO(file_content), 
                    read_only=True,
                    data_only=True,
                    keep_links=False
                )
                
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    return None
                
                worksheet = workbook[sheet_name]
                
                # 读取数据 - 限制扫描范围
                data_rows = []
                headers = []
                
                # 读取表头 - 限制列数
                max_col = min(worksheet.max_column or 0, self.max_cols_to_check)
                for col_num in range(1, max_col + 1):
                    try:
                        cell_value = worksheet.cell(1, col_num).value
                        header = str(cell_value).strip() if cell_value is not None else f"列{col_num}"
                        headers.append(header)
                    except Exception:
                        headers.append(f"列{col_num}")
                
                # 读取数据 - 限制行数
                max_row = min(worksheet.max_row or 0, self.max_rows_to_scan)
                for row_num in range(2, max_row + 1):
                    row_data = {}
                    has_data = False
                    
                    for col_idx, header in enumerate(headers):
                        col_num = col_idx + 1
                        try:
                            cell_value = worksheet.cell(row_num, col_num).value
                            value = str(cell_value).strip() if cell_value is not None else ""
                            row_data[header] = value
                            if value:
                                has_data = True
                        except Exception:
                            row_data[header] = ""
                    
                    if has_data:
                        row_data['_row_index'] = row_num - 2  # 相对于数据行的索引
                        data_rows.append(row_data)
                
                workbook.close()
                gc.collect()  # 强制垃圾回收
                
                cached_data = {
                    'headers': headers,
                    'data_rows': data_rows
                }
                
                # 缓存数据 - 但设置较短的过期时间以节省内存
                self._add_to_cache(cache_key, cached_data, timeout=300)  # 5分钟过期
            
            # 在缓存的数据中搜索
            return self._search_in_data(cached_data, search_code, fuzzy_match)
            
        except Exception as e:
            st.error(f"在工作表中搜索失败: {str(e)}")
            return None
    
    def _search_in_data(self, cached_data: Dict[str, Any], search_code: str, fuzzy_match: bool) -> Dict[str, Any]:
        """在缓存的数据中执行搜索 - 性能优化版本"""
        headers = cached_data['headers']
        data_rows = cached_data['data_rows']
        
        matches = []
        search_pattern = search_code.strip().lower()
        
        # 限制搜索结果数量以提高性能
        max_matches = 100
        
        for row_data in data_rows:
            if len(matches) >= max_matches:
                break
                
            row_index = row_data.get('_row_index', 0)
            
            for header in headers:
                if header == '_row_index':
                    continue
                
                cell_value = str(row_data.get(header, '')).strip()
                
                if cell_value:
                    match_found = False
                    if fuzzy_match:
                        # 模糊匹配：包含搜索内容
                        if search_pattern in cell_value.lower():
                            match_found = True
                    else:
                        # 精确匹配
                        if search_pattern == cell_value.lower():
                            match_found = True
                    
                    if match_found:
                        match_data = {k: v for k, v in row_data.items() if k != '_row_index'}
                        matches.append({
                            'row_index': row_index,
                            'column': header,
                            'matched_value': cell_value,
                            'row_data': match_data
                        })
                        break  # 每行只匹配一次，提高性能
        
        return {
            'match_count': len(matches),
            'matches': matches,
            'truncated': len(matches) >= max_matches
        }
    
    def _get_from_cache(self, key: str) -> Optional[Any]:
        """从缓存获取数据 - 内存优化版本"""
        if key in self.sheet_cache:
            cache_entry = self.sheet_cache[key]
            timestamp = cache_entry.get('timestamp', 0)
            timeout = cache_entry.get('timeout', self.cache_timeout)
            
            # 检查是否过期
            if (datetime.now().timestamp() - timestamp) < timeout:
                return cache_entry.get('data')
            else:
                # 删除过期缓存
                del self.sheet_cache[key]
        
        return None
    
    def _add_to_cache(self, key: str, data: Any, timeout: Optional[int] = None):
        """添加数据到缓存 - 内存优化版本"""
        # 如果缓存已满，删除最旧的条目
        if len(self.sheet_cache) >= self.max_cache_size:
            # 删除一半的旧缓存以减少频繁清理
            sorted_keys = sorted(
                self.sheet_cache.keys(),
                key=lambda k: self.sheet_cache[k].get('timestamp', 0)
            )
            for old_key in sorted_keys[:self.max_cache_size // 2]:
                del self.sheet_cache[old_key]
        
        self.sheet_cache[key] = {
            'data': data,
            'timestamp': datetime.now().timestamp(),
            'timeout': timeout or self.cache_timeout
        }
    
    def get_cache_info(self) -> Dict[str, Any]:
        """获取缓存信息"""
        cached_sheets = list(self.sheet_cache.keys())
        cache_size_mb = sum(
            len(str(entry.get('data', ''))) for entry in self.sheet_cache.values()
        ) / (1024 * 1024)
        
        return {
            'sheet_data_cache_size': len(self.sheet_cache),
            'max_cache_size': self.max_cache_size,
            'cached_sheets': cached_sheets,
            'cache_timeout': self.cache_timeout,
            'estimated_cache_size_mb': round(cache_size_mb, 2)
        }
    
    def clear_cache(self):
        """清除所有缓存"""
        self.sheet_cache.clear()
        gc.collect()  # 强制垃圾回收
    
    def optimize_memory(self):
        """优化内存使用 - 增强版本"""
        current_time = datetime.now().timestamp()
        
        # 删除过期缓存
        expired_keys = []
        for key, cache_entry in self.sheet_cache.items():
            timeout = cache_entry.get('timeout', self.cache_timeout)
            if (current_time - cache_entry.get('timestamp', 0)) > timeout:
                expired_keys.append(key)
        
        for key in expired_keys:
            del self.sheet_cache[key]
        
        # 强制垃圾回收
        gc.collect()
        
        st.info(f"已清理 {len(expired_keys)} 个过期缓存条目，释放内存")
    
    def get_memory_usage(self) -> Dict[str, Any]:
        """获取内存使用情况"""
        import sys
        
        cache_memory = 0
        for entry in self.sheet_cache.values():
            try:
                cache_memory += sys.getsizeof(entry.get('data', ''))
            except Exception:
                pass
        
        return {
            'cache_entries': len(self.sheet_cache),
            'cache_memory_mb': round(cache_memory / (1024 * 1024), 2),
            'max_file_size_mb': self.max_file_size // (1024 * 1024)
        }
    
    def get_file_statistics(self, file_content: bytes) -> Dict[str, Any]:
        """获取Excel文件统计信息"""
        try:
            workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            
            sheet_names = workbook.sheetnames
            total_sheets = len(sheet_names)
            file_size = len(file_content)
            
            sheets_info = []
            valid_sheet_names = []
            
            for sheet_name in sheet_names:
                try:
                    worksheet = workbook[sheet_name]
                    
                    # 计算实际使用的行列数
                    max_row = worksheet.max_row or 0
                    max_col = worksheet.max_column or 0
                    
                    # 检查是否有数据（排除只有标题行的情况）
                    has_data = max_row > 1 and max_col > 0
                    
                    # 如果有数据，进一步验证
                    if has_data:
                        # 检查前几行是否有实际内容
                        data_found = False
                        for row_num in range(1, min(6, max_row + 1)):  # 检查前5行
                            row_values = []
                            for col_num in range(1, min(max_col + 1, 20)):  # 检查前20列
                                cell_value = worksheet.cell(row_num, col_num).value
                                if cell_value is not None and str(cell_value).strip():
                                    row_values.append(str(cell_value).strip())
                            
                            if len(row_values) > 0:
                                data_found = True
                                break
                        
                        has_data = data_found
                    
                    sheet_info = {
                        'name': sheet_name,
                        'rows': max_row,
                        'columns': max_col,
                        'has_data': has_data
                    }
                    
                    sheets_info.append(sheet_info)
                    
                    # 只有有数据的工作表才算作有效门店
                    if has_data:
                        valid_sheet_names.append(sheet_name)
                        
                except Exception as e:
                    st.warning(f"分析工作表 '{sheet_name}' 时出错: {str(e)}")
                    sheets_info.append({
                        'name': sheet_name,
                        'rows': 0,
                        'columns': 0,
                        'has_data': False
                    })
            
            workbook.close()
            
            return {
                'total_sheets': total_sheets,
                'file_size': file_size,
                'sheets_info': sheets_info,
                'sheet_names': valid_sheet_names,  # 只返回有数据的工作表
                'valid_sheets': len(valid_sheet_names)
            }
            
        except Exception as e:
            st.error(f"分析Excel文件失败: {str(e)}")
            return {
                'total_sheets': 0,
                'file_size': len(file_content),
                'sheets_info': [],
                'sheet_names': [],
                'valid_sheets': 0
            }
    
    def get_sheet_names_fast(self, file_content: bytes) -> List[str]:
        """快速获取有效工作表名称"""
        try:
            stats = self.get_file_statistics(file_content)
            return stats.get('sheet_names', [])
        except Exception as e:
            st.error(f"获取工作表名称失败: {str(e)}")
            return []
    
    def get_sheet_preview(self, file_content: bytes, sheet_name: str, limit: int = 5) -> Optional[Dict[str, Any]]:
        """获取工作表预览数据"""
        try:
            # 检查缓存
            cache_key = f"{sheet_name}_preview_{limit}"
            cached_data = self._get_from_cache(cache_key)
            if cached_data:
                return cached_data
            
            # 读取Excel文件
            workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return None
            
            worksheet = workbook[sheet_name]
            
            # 获取基本信息
            total_rows = worksheet.max_row or 0
            total_columns = worksheet.max_column or 0
            
            # 读取预览数据
            preview_data = []
            headers = []
            
            # 读取表头（第一行）
            if total_rows > 0:
                for col_num in range(1, min(total_columns + 1, 20)):  # 最多20列
                    cell_value = worksheet.cell(1, col_num).value
                    header = str(cell_value).strip() if cell_value is not None else f"列{col_num}"
                    headers.append(header)
            
            # 读取数据行
            for row_num in range(2, min(total_rows + 1, limit + 2)):  # 跳过表头
                row_data = {}
                has_data = False
                
                for col_idx, header in enumerate(headers):
                    col_num = col_idx + 1
                    cell_value = worksheet.cell(row_num, col_num).value
                    value = str(cell_value).strip() if cell_value is not None else ""
                    row_data[header] = value
                    if value:
                        has_data = True
                
                if has_data:
                    preview_data.append(row_data)
            
            workbook.close()
            
            result = {
                'total_rows': total_rows,
                'total_columns': total_columns,
                'preview_data': preview_data,
                'headers': headers
            }
            
            # 缓存结果
            self._add_to_cache(cache_key, result)
            
            return result
            
        except Exception as e:
            st.error(f"获取工作表预览失败: {str(e)}")
            return None
    
    def search_in_sheet(self, file_content: bytes, sheet_name: str, search_code: str, fuzzy_match: bool = True) -> Optional[Dict[str, Any]]:
        """在指定工作表中搜索编码"""
        try:
            # 检查缓存
            cache_key = f"{sheet_name}_data"
            cached_data = self._get_from_cache(cache_key)
            
            if not cached_data:
                # 读取Excel文件
                workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    return None
                
                worksheet = workbook[sheet_name]
                
                # 读取数据到DataFrame
                data_rows = []
                headers = []
                
                # 读取表头
                max_col = min(worksheet.max_column or 0, 50)  # 限制最大列数
                for col_num in range(1, max_col + 1):
                    cell_value = worksheet.cell(1, col_num).value
                    header = str(cell_value).strip() if cell_value is not None else f"列{col_num}"
                    headers.append(header)
                
                # 读取数据
                max_row = min(worksheet.max_row or 0, self.max_rows_to_scan)
                for row_num in range(2, max_row + 1):
                    row_data = {}
                    has_data = False
                    
                    for col_idx, header in enumerate(headers):
                        col_num = col_idx + 1
                        cell_value = worksheet.cell(row_num, col_num).value
                        value = str(cell_value).strip() if cell_value is not None else ""
                        row_data[header] = value
                        if value:
                            has_data = True
                    
                    if has_data:
                        row_data['_row_index'] = row_num - 2  # 相对于数据行的索引
                        data_rows.append(row_data)
                
                workbook.close()
                
                cached_data = {
                    'headers': headers,
                    'data_rows': data_rows
                }
                
                # 缓存数据
                self._add_to_cache(cache_key, cached_data)
            
            # 在缓存的数据中搜索
            return self._search_in_data(cached_data, search_code, fuzzy_match)
            
        except Exception as e:
            st.error(f"在工作表中搜索失败: {str(e)}")
            return None
    
    def _search_in_data(self, cached_data: Dict[str, Any], search_code: str, fuzzy_match: bool) -> Dict[str, Any]:
        """在缓存的数据中执行搜索"""
        headers = cached_data['headers']
        data_rows = cached_data['data_rows']
        
        matches = []
        search_pattern = search_code.strip().lower()
        
        for row_data in data_rows:
            row_index = row_data.get('_row_index', 0)
            
            for header in headers:
                if header == '_row_index':
                    continue
                
                cell_value = str(row_data.get(header, '')).strip()
                
                if cell_value:
                    if fuzzy_match:
                        # 模糊匹配：包含搜索内容
                        if search_pattern in cell_value.lower():
                            match_data = {k: v for k, v in row_data.items() if k != '_row_index'}
                            matches.append({
                                'row_index': row_index,
                                'column': header,
                                'matched_value': cell_value,
                                'row_data': match_data
                            })
                    else:
                        # 精确匹配
                        if search_pattern == cell_value.lower():
                            match_data = {k: v for k, v in row_data.items() if k != '_row_index'}
                            matches.append({
                                'row_index': row_index,
                                'column': header,
                                'matched_value': cell_value,
                                'row_data': match_data
                            })
        
        return {
            'match_count': len(matches),
            'matches': matches
        }
    
    def _get_from_cache(self, key: str) -> Optional[Any]:
        """从缓存获取数据"""
        if key in self.sheet_cache:
            cache_entry = self.sheet_cache[key]
            timestamp = cache_entry.get('timestamp', 0)
            
            # 检查是否过期
            if (datetime.now().timestamp() - timestamp) < self.cache_timeout:
                return cache_entry.get('data')
            else:
                # 删除过期缓存
                del self.sheet_cache[key]
        
        return None
    
    def _add_to_cache(self, key: str, data: Any):
        """添加数据到缓存"""
        # 如果缓存已满，删除最旧的条目
        if len(self.sheet_cache) >= self.max_cache_size:
            oldest_key = min(
                self.sheet_cache.keys(),
                key=lambda k: self.sheet_cache[k].get('timestamp', 0)
            )
            del self.sheet_cache[oldest_key]
        
        self.sheet_cache[key] = {
            'data': data,
            'timestamp': datetime.now().timestamp()
        }
    
    def get_cache_info(self) -> Dict[str, Any]:
        """获取缓存信息"""
        cached_sheets = list(self.sheet_cache.keys())
        
        return {
            'sheet_data_cache_size': len(self.sheet_cache),
            'max_cache_size': self.max_cache_size,
            'cached_sheets': cached_sheets,
            'cache_timeout': self.cache_timeout
        }
    
    def clear_cache(self):
        """清除所有缓存"""
        self.sheet_cache.clear()
    
    def clear_sheet_cache(self, sheet_name: str):
        """清除指定工作表的缓存"""
        keys_to_remove = [key for key in self.sheet_cache.keys() if sheet_name in key]
        for key in keys_to_remove:
            del self.sheet_cache[key]
    
    def optimize_memory(self):
        """优化内存使用"""
        current_time = datetime.now().timestamp()
        
        # 删除过期缓存
        expired_keys = []
        for key, cache_entry in self.sheet_cache.items():
            if (current_time - cache_entry.get('timestamp', 0)) > self.cache_timeout:
                expired_keys.append(key)
        
        for key in expired_keys:
            del self.sheet_cache[key]
        
        st.info(f"已清理 {len(expired_keys)} 个过期缓存条目")
    
    def get_sheet_analysis(self, file_content: bytes, sheet_name: str) -> Optional[Dict[str, Any]]:
        """分析工作表数据结构"""
        try:
            preview_data = self.get_sheet_preview(file_content, sheet_name, 100)  # 分析更多行
            
            if not preview_data:
                return None
            
            headers = preview_data.get('headers', [])
            data_rows = preview_data.get('preview_data', [])
            
            # 分析每列的数据类型和特征
            column_analysis = {}
            for header in headers:
                values = [row.get(header, '') for row in data_rows if row.get(header, '').strip()]
                
                # 数据类型分析
                numeric_count = 0
                date_count = 0
                text_count = 0
                
                for value in values:
                    if value.strip():
                        if self._is_numeric(value):
                            numeric_count += 1
                        elif self._is_date(value):
                            date_count += 1
                        else:
                            text_count += 1
                
                total_values = len(values)
                column_analysis[header] = {
                    'total_values': total_values,
                    'numeric_ratio': numeric_count / total_values if total_values > 0 else 0,
                    'date_ratio': date_count / total_values if total_values > 0 else 0,
                    'text_ratio': text_count / total_values if total_values > 0 else 0,
                    'sample_values': values[:5]  # 前5个示例值
                }
            
            return {
                'total_rows': preview_data.get('total_rows', 0),
                'total_columns': preview_data.get('total_columns', 0),
                'headers': headers,
                'column_analysis': column_analysis,
                'data_quality': self._assess_data_quality(data_rows, headers)
            }
            
        except Exception as e:
            st.error(f"分析工作表结构失败: {str(e)}")
            return None
    
    def _is_numeric(self, value: str) -> bool:
        """检查值是否为数字"""
        try:
            float(value.replace(',', ''))  # 支持千分位分隔符
            return True
        except ValueError:
            return False
    
    def _is_date(self, value: str) -> bool:
        """检查值是否为日期"""
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'\d{4}年\d{1,2}月\d{1,2}日'
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value.strip()):
                return True
        return False
    
    def _assess_data_quality(self, data_rows: List[Dict], headers: List[str]) -> Dict[str, Any]:
        """评估数据质量"""
        if not data_rows:
            return {'score': 0, 'issues': ['无数据']}
        
        total_cells = len(data_rows) * len(headers)
        empty_cells = 0
        issues = []
        
        for row in data_rows:
            for header in headers:
                if not row.get(header, '').strip():
                    empty_cells += 1
        
        fill_rate = (total_cells - empty_cells) / total_cells if total_cells > 0 else 0
        
        if fill_rate < 0.5:
            issues.append('数据填充率较低')
        
        if len(headers) < 3:
            issues.append('列数较少')
        
        if len(data_rows) < 10:
            issues.append('数据行数较少')
        
        # 计算质量分数 (0-100)
        quality_score = int(fill_rate * 100)
        
        return {
            'score': quality_score,
            'fill_rate': fill_rate,
            'total_cells': total_cells,
            'empty_cells': empty_cells,
            'issues': issues
        }
