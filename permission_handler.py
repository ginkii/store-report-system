import pandas as pd
import streamlit as st
import io
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
import gc  # 垃圾回收

class PermissionHandler:
    def __init__(self):
        # 文件大小限制
        self.max_file_size = 5 * 1024 * 1024  # 5MB，权限表通常较小
        
        # 导入时动态选择存储处理器
        try:
            from cos_handler import COSHandler
            self.storage_handler = COSHandler()
            self.storage_type = "COS"
        except ImportError:
            from local_storage_handler import LocalStorageHandler
            self.storage_handler = LocalStorageHandler()
            self.storage_type = "LOCAL"
        
        from json_handler import JSONHandler
        self.json_handler = JSONHandler()
    
    def validate_permission_file(self, file_content: bytes, progress_callback=None) -> Tuple[bool, str]:
        """验证权限表文件格式 - openpyxl优化版本"""
        try:
            if progress_callback:
                progress_callback(10, "正在验证文件格式...")
            
            # 文件大小检查
            if len(file_content) > self.max_file_size:
                return False, f"文件大小超过限制 ({self.max_file_size // (1024*1024)}MB)"
            
            # 文件格式快速验证
            if len(file_content) < 1024:
                return False, "文件太小，可能不是有效的Excel文件"
            
            if progress_callback:
                progress_callback(30, "正在打开Excel文件...")
            
            # 使用内存优化的只读模式
            try:
                workbook = load_workbook(
                    io.BytesIO(file_content), 
                    read_only=True,
                    data_only=True,
                    keep_links=False
                )
                
                if progress_callback:
                    progress_callback(50, "正在检查工作表结构...")
                
                # 基本检查
                if len(workbook.sheetnames) == 0:
                    workbook.close()
                    return False, "Excel文件中没有工作表"
                
                # 获取第一个工作表
                worksheet = workbook[workbook.sheetnames[0]]
                
                if progress_callback:
                    progress_callback(70, "正在验证数据结构...")
                
                # 检查工作表是否有数据
                max_row = worksheet.max_row or 0
                max_col = worksheet.max_column or 0
                
                if max_row < 2:  # 至少需要标题行+数据行
                    workbook.close()
                    return False, "文件中没有足够的数据行（至少需要2行：标题+数据）"
                
                if max_col < 2:  # 至少需要门店名称和查询编码两列
                    workbook.close()
                    return False, "文件必须包含至少两列数据（门店名称和查询编码）"
                
                if progress_callback:
                    progress_callback(90, "正在验证数据内容...")
                
                # 快速检查前几行数据
                has_valid_data = False
                for row_num in range(2, min(6, max_row + 1)):  # 检查前5行数据
                    store_cell = worksheet.cell(row_num, 1).value
                    code_cell = worksheet.cell(row_num, 2).value
                    
                    if store_cell is not None and code_cell is not None:
                        store_value = str(store_cell).strip()
                        code_value = str(code_cell).strip()
                        
                        if store_value and code_value and store_value != 'None' and code_value != 'None':
                            has_valid_data = True
                            break
                
                workbook.close()
                gc.collect()
                
                if not has_valid_data:
                    return False, "文件中没有找到有效的权限数据"
                
                if progress_callback:
                    progress_callback(100, "文件验证通过")
                
                return True, "验证通过"
                
            except openpyxl.utils.exceptions.InvalidFileException:
                return False, "不是有效的Excel文件格式"
            except Exception as e:
                return False, f"文件格式验证失败: {str(e)}"
                
        except Exception as e:
            if progress_callback:
                progress_callback(0, f"验证失败: {str(e)}")
            return False, f"文件验证过程中出错: {str(e)}"
    
    def parse_permission_file(self, file_content: bytes, progress_callback=None) -> Tuple[bool, List[Dict[str, str]], str]:
        """解析权限表文件 - openpyxl优化版本"""
        try:
            if progress_callback:
                progress_callback(10, "正在打开权限表文件...")
            
            # 内存优化的文件加载
            workbook = load_workbook(
                io.BytesIO(file_content), 
                read_only=True,
                data_only=True,
                keep_links=False
            )
            
            # 获取第一个工作表
            worksheet = workbook[workbook.sheetnames[0]]
            max_row = worksheet.max_row or 0
            
            if progress_callback:
                progress_callback(30, f"开始解析数据，共 {max_row-1} 行...")
            
            permissions = []
            processed_count = 0
            valid_count = 0
            
            # 跳过第一行（标题行），从第二行开始解析
            for row_num in range(2, max_row + 1):
                try:
                    if progress_callback and processed_count % 10 == 0:  # 每10行更新一次进度
                        progress = 30 + int((processed_count / (max_row - 1)) * 60)
                        progress_callback(progress, f"正在解析第 {processed_count+1} 行...")
                    
                    # 读取门店名称（第一列）和查询编码（第二列）
                    store_cell = worksheet.cell(row_num, 1).value
                    code_cell = worksheet.cell(row_num, 2).value
                    
                    # 数据清理和验证
                    if store_cell is not None and code_cell is not None:
                        store_name = str(store_cell).strip()
                        query_code = str(code_cell).strip()
                        
                        # 跳过空值和无效值
                        if (store_name and query_code and 
                            store_name.lower() not in ['none', 'null', 'nan', ''] and
                            query_code.lower() not in ['none', 'null', 'nan', '']):
                            
                            permissions.append({
                                'store': store_name,
                                'code': query_code
                            })
                            valid_count += 1
                    
                    processed_count += 1
                    
                except Exception as e:
                    # 单行解析失败，记录警告但继续处理
                    st.warning(f"解析第 {row_num} 行时出错: {str(e)}")
                    processed_count += 1
                    continue
            
            workbook.close()
            gc.collect()  # 强制垃圾回收
            
            if progress_callback:
                progress_callback(100, f"解析完成！共处理 {processed_count} 行，有效记录 {valid_count} 条")
            
            if not permissions:
                return False, [], "解析后没有有效的权限记录"
            
            return True, permissions, f"成功解析 {len(permissions)} 条权限记录"
            
        except Exception as e:
            if progress_callback:
                progress_callback(0, f"解析失败: {str(e)}")
            return False, [], f"解析失败: {str(e)}"
    
    def get_file_statistics(self, file_content: bytes, progress_callback=None) -> Dict[str, Any]:
        """获取权限表文件统计信息 - openpyxl优化版本"""
        try:
            if progress_callback:
                progress_callback(10, "正在分析文件统计...")
            
            # 内存优化的文件加载
            workbook = load_workbook(
                io.BytesIO(file_content), 
                read_only=True,
                data_only=True,
                keep_links=False
            )
            
            # 获取第一个工作表
            worksheet = workbook[workbook.sheetnames[0]]
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0
            
            if progress_callback:
                progress_callback(40, "正在统计数据...")
            
            # 基本统计
            total_rows = max_row - 1 if max_row > 0 else 0  # 减去标题行
            
            # 读取表头
            headers = []
            if max_row > 0:
                for col_num in range(1, min(max_col + 1, 10)):  # 最多读取10列表头
                    cell_value = worksheet.cell(1, col_num).value
                    header = str(cell_value).strip() if cell_value is not None else f"列{col_num}"
                    headers.append(header)
            
            if progress_callback:
                progress_callback(70, "正在分析数据质量...")
            
            # 快速扫描数据质量（只检查前50行以提高性能）
            stores_set = set()
            codes_set = set()
            valid_records = 0
            scan_limit = min(max_row, 52)  # 标题行+最多50行数据
            
            for row_num in range(2, scan_limit + 1):
                try:
                    store_cell = worksheet.cell(row_num, 1).value
                    code_cell = worksheet.cell(row_num, 2).value
                    
                    if store_cell is not None and code_cell is not None:
                        store_name = str(store_cell).strip()
                        query_code = str(code_cell).strip()
                        
                        if (store_name and query_code and 
                            store_name.lower() not in ['none', 'null', 'nan', ''] and
                            query_code.lower() not in ['none', 'null', 'nan', '']):
                            
                            stores_set.add(store_name)
                            codes_set.add(query_code)
                            valid_records += 1
                            
                except Exception:
                    continue
            
            workbook.close()
            gc.collect()
            
            if progress_callback:
                progress_callback(100, "统计分析完成")
            
            # 如果扫描了全部数据，直接返回；否则按比例估算
            if scan_limit >= max_row:
                estimated_unique_stores = len(stores_set)
                estimated_unique_codes = len(codes_set)
                estimated_valid_records = valid_records
            else:
                # 按比例估算（保守估计）
                scan_ratio = (scan_limit - 1) / total_rows if total_rows > 0 else 1
                estimated_unique_stores = int(len(stores_set) / scan_ratio * 0.8)  # 保守估计
                estimated_unique_codes = int(len(codes_set) / scan_ratio * 0.8)
                estimated_valid_records = int(valid_records / scan_ratio * 0.9)
            
            return {
                'total_rows': total_rows,
                'valid_records': estimated_valid_records,
                'unique_stores': estimated_unique_stores,
                'unique_codes': estimated_unique_codes,
                'file_size': len(file_content),
                'columns': headers,
                'scanned_rows': min(50, total_rows),
                'is_estimated': scan_limit < max_row
            }
            
        except Exception as e:
            if progress_callback:
                progress_callback(0, f"统计分析失败: {str(e)}")
            st.error(f"获取文件统计失败: {str(e)}")
            return {
                'total_rows': 0,
                'valid_records': 0,
                'unique_stores': 0,
                'unique_codes': 0,
                'file_size': len(file_content),
                'columns': [],
                'scanned_rows': 0,
                'is_estimated': False
            }
    
    def upload_permission_file(self, file_content: bytes, file_name: str, progress_callback=None) -> Optional[str]:
        """上传权限表文件 - 支持进度显示"""
        try:
            if progress_callback:
                progress_callback(10, "准备上传权限表文件...")
            
            # 构造存储路径
            folder = "permissions"
            
            # 检查存储处理器是否支持进度回调
            try:
                import inspect
                sig = inspect.signature(self.storage_handler.upload_file)
                params = list(sig.parameters.keys())
                
                if 'progress_callback' in params or len(params) > 4:
                    def upload_progress_callback(percent, message):
                        # 上传占用 20-90% 的进度
                        adjusted_percent = 20 + int(percent * 0.7)
                        if progress_callback:
                            progress_callback(adjusted_percent, f"上传中: {message}")
                    
                    file_path = self.storage_handler.upload_file(
                        file_content, file_name, folder, 
                        progress_callback=upload_progress_callback
                    )
                else:
                    file_path = self.storage_handler.upload_file(file_content, file_name, folder)
                    if progress_callback:
                        progress_callback(80, "文件上传完成")
            except:
                # 兜底方案：不使用进度回调
                file_path = self.storage_handler.upload_file(file_content, file_name, folder)
                if progress_callback:
                    progress_callback(80, "文件上传完成")
            
            if file_path:
                if progress_callback:
                    progress_callback(100, "权限表文件上传成功")
                st.success(f"权限表文件上传成功: {file_path}")
                return file_path
            else:
                if progress_callback:
                    progress_callback(0, "文件上传失败")
                st.error("权限表文件上传失败")
                return None
                
        except Exception as e:
            if progress_callback:
                progress_callback(0, f"上传失败: {str(e)}")
            st.error(f"上传权限表文件失败: {str(e)}")
            return None
    
    def update_permissions(self, file_path: str, permissions: List[Dict[str, str]], 
                          file_name: str, file_size: int) -> bool:
        """更新权限表数据"""
        try:
            # 构造权限表信息
            permission_info = {
                'file_info': {
                    'file_name': file_name,
                    'file_path': file_path,
                    'upload_time': datetime.now().isoformat(),
                    'file_size': file_size,
                    'total_records': len(permissions),
                    'storage_type': self.storage_type
                },
                'permissions': permissions
            }
            
            # 保存到JSON
            return self.json_handler.update_permissions(permission_info)
            
        except Exception as e:
            st.error(f"更新权限数据失败: {str(e)}")
            return False
    
    def get_current_permissions(self) -> Optional[Dict[str, Any]]:
        """获取当前权限表信息"""
        try:
            return self.json_handler.get_permissions()
        except Exception as e:
            st.error(f"获取权限信息失败: {str(e)}")
            return None
    
    def check_permission(self, store_name: str, query_code: str) -> bool:
        """检查门店-编码组合是否有权限"""
        try:
            permissions_data = self.get_current_permissions()
            if not permissions_data:
                return False
            
            permissions = permissions_data.get('permissions', [])
            
            # 检查是否存在匹配的权限记录
            for permission in permissions:
                if permission.get('store') == store_name and permission.get('code') == query_code:
                    return True
            
            return False
            
        except Exception as e:
            st.error(f"权限检查失败: {str(e)}")
            return False
    
    def get_permission_statistics(self) -> Dict[str, Any]:
        """获取权限统计信息"""
        try:
            permissions_data = self.get_current_permissions()
            if not permissions_data:
                return {
                    'total_records': 0,
                    'unique_stores': 0,
                    'unique_codes': 0,
                    'file_info': None,
                    'has_permissions': False
                }
            
            permissions = permissions_data.get('permissions', [])
            file_info = permissions_data.get('file_info', {})
            
            if not permissions:
                return {
                    'total_records': 0,
                    'unique_stores': 0,
                    'unique_codes': 0,
                    'file_info': file_info,
                    'has_permissions': False
                }
            
            # 统计唯一门店和编码
            stores = set(p.get('store') for p in permissions)
            codes = set(p.get('code') for p in permissions)
            
            return {
                'total_records': len(permissions),
                'unique_stores': len(stores),
                'unique_codes': len(codes),
                'file_info': file_info,
                'has_permissions': True,
                'stores': list(stores),
                'codes': list(codes)
            }
            
        except Exception as e:
            st.error(f"获取权限统计失败: {str(e)}")
            return {
                'total_records': 0,
                'unique_stores': 0,
                'unique_codes': 0,
                'file_info': None,
                'has_permissions': False
            }
    
    def get_permissions_preview(self, limit: int = 10) -> List[Dict[str, str]]:
        """获取权限记录预览"""
        try:
            permissions_data = self.get_current_permissions()
            if not permissions_data:
                return []
            
            permissions = permissions_data.get('permissions', [])
            return permissions[:limit]
            
        except Exception as e:
            st.error(f"获取权限预览失败: {str(e)}")
            return []
    
    def validate_permissions_with_stores(self, available_stores: List[str]) -> Dict[str, Any]:
        """验证权限表中的门店是否在汇总报表中存在"""
        try:
            permissions_data = self.get_current_permissions()
            if not permissions_data:
                return {
                    'valid': True,
                    'invalid_stores': [],
                    'orphaned_permissions': 0,
                    'total_permission_stores': 0,
                    'available_stores': len(available_stores)
                }
            
            permissions = permissions_data.get('permissions', [])
            
            # 检查权限表中的门店
            permission_stores = set(p.get('store') for p in permissions)
            available_stores_set = set(available_stores)
            
            # 找出不在汇总报表中的门店
            invalid_stores = permission_stores - available_stores_set
            
            # 统计孤立的权限记录
            orphaned_count = sum(1 for p in permissions 
                               if p.get('store') in invalid_stores)
            
            return {
                'valid': len(invalid_stores) == 0,
                'invalid_stores': list(invalid_stores),
                'orphaned_permissions': orphaned_count,
                'total_permission_stores': len(permission_stores),
                'available_stores': len(available_stores_set)
            }
            
        except Exception as e:
            st.error(f"权限验证失败: {str(e)}")
            return {
                'valid': False,
                'invalid_stores': [],
                'orphaned_permissions': 0,
                'total_permission_stores': 0,
                'available_stores': len(available_stores)
            }
    
    def clear_permissions(self) -> bool:
        """清空权限表"""
        try:
            return self.json_handler.clear_permissions()
        except Exception as e:
            st.error(f"清空权限失败: {str(e)}")
            return False
    
    def export_permissions(self) -> Optional[bytes]:
        """导出权限表为Excel - 使用openpyxl优化"""
        try:
            permissions_data = self.get_current_permissions()
            if not permissions_data:
                return None
            
            permissions = permissions_data.get('permissions', [])
            if not permissions:
                return None
            
            # 使用openpyxl直接创建Excel文件以提高性能
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "权限表"
            
            # 设置表头
            headers = ['门店名称', '查询编码']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 写入数据
            for row_num, permission in enumerate(permissions, 2):
                ws.cell(row=row_num, column=1, value=permission.get('store', ''))
                ws.cell(row=row_num, column=2, value=permission.get('code', ''))
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存为字节流
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"导出权限表失败: {str(e)}")
            return None
